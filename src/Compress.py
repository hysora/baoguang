# -*- coding: utf-8 -*-
"""
下载 avatar_over_2m.xlsx 里大于 2MB 的头像，统一压缩成 WebP（≤2MB），存到桌面。

策略：
  - 长边限制到 1080px（只缩小、不放大，保持比例）
  - 统一转 WebP；静态图转静态 WebP，动图（GIF/WebP 动画）转动态 WebP
  - 自适应：先 quality 85，若仍 >2MB 依次降 quality，再不行继续缩尺寸，直到达标
  - 保留透明通道；动图保留每帧时长与循环

依赖：
    pip install openpyxl requests Pillow
"""
import os
import io
import threading
import concurrent.futures

import requests
from openpyxl import load_workbook, Workbook
from PIL import Image, ImageSequence

# ---------------- 配置 ----------------
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
SRC_LIST = os.path.join(DESKTOP, "avatar_over_2m.xlsx")   # Test.py 产出的 >2M 列表
OUT_DIR = os.path.join(DESKTOP, "avatar_compressed")       # 压缩后图片输出目录
REPORT = os.path.join(DESKTOP, "avatar_compressed_report.xlsx")

SIZE_LIMIT = 2 * 1024 * 1024   # 目标：<2MB
MAX_EDGE = 1080                # 长边上限(px)
START_QUALITY = 85             # 初始 WebP 质量
MIN_QUALITY = 50              # 质量下限，到此还超标就缩尺寸
MAX_WORKERS = 12              # 并发下载+压缩
TIMEOUT = 30

_session = requests.Session()
_session.headers.update({"User-Agent": "Mozilla/5.0"})
_print_lock = threading.Lock()


def log(msg):
    with _print_lock:
        print(msg)


def download(url):
    """下载原图字节，失败返回 None。"""
    try:
        resp = _session.get(url, timeout=TIMEOUT)
        resp.raise_for_status()
        return resp.content
    except Exception as e:
        log(f"[WARN] 下载失败 {url} -> {e}")
        return None


def _scaled_size(w, h, max_edge):
    """按长边上限等比缩放后的尺寸（不放大）。"""
    if max(w, h) <= max_edge:
        return w, h
    if w >= h:
        return max_edge, max(1, round(h * max_edge / w))
    return max(1, round(w * max_edge / h)), max_edge


def _has_alpha(img):
    return img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info)


def compress_static(img, max_edge):
    """压缩静态图为 WebP 字节，自适应直到 <2MB。返回 (bytes, final_quality, final_edge)。"""
    if _has_alpha(img):
        base = img.convert("RGBA")
    else:
        base = img.convert("RGB")

    edge = max_edge
    while True:
        w, h = base.size
        nw, nh = _scaled_size(w, h, edge)
        frame = base.resize((nw, nh), Image.LANCZOS) if (nw, nh) != (w, h) else base

        quality = START_QUALITY
        while quality >= MIN_QUALITY:
            buf = io.BytesIO()
            frame.save(buf, format="WEBP", quality=quality, method=6)
            data = buf.getvalue()
            if len(data) <= SIZE_LIMIT:
                return data, quality, edge
            quality -= 10

        # 质量已到下限仍超标 -> 缩尺寸再来
        if edge <= 256:
            # 实在压不下去，返回当前最小结果（最低质量+最小尺寸）
            return data, MIN_QUALITY, edge
        edge = int(edge * 0.8)


def compress_animated(img, max_edge):
    """压缩动图为动态 WebP 字节，自适应直到 <2MB。返回 (bytes, final_quality, final_edge)。"""
    # 预读所有帧及时长
    raw_frames = []
    durations = []
    for frame in ImageSequence.Iterator(img):
        raw_frames.append(frame.convert("RGBA"))
        durations.append(frame.info.get("duration", 100))
    loop = img.info.get("loop", 0)

    w, h = raw_frames[0].size
    edge = max_edge
    while True:
        nw, nh = _scaled_size(w, h, edge)
        if (nw, nh) != (w, h):
            frames = [f.resize((nw, nh), Image.LANCZOS) for f in raw_frames]
        else:
            frames = raw_frames

        quality = START_QUALITY
        while quality >= MIN_QUALITY:
            buf = io.BytesIO()
            frames[0].save(
                buf, format="WEBP", save_all=True, append_images=frames[1:],
                duration=durations, loop=loop, quality=quality, method=4,
            )
            data = buf.getvalue()
            if len(data) <= SIZE_LIMIT:
                return data, quality, edge
            quality -= 10

        if edge <= 256:
            return data, MIN_QUALITY, edge
        edge = int(edge * 0.8)


def process(uid, url):
    """下载并压缩单个头像，返回报表行 dict。"""
    row = {"uid": uid, "url": url, "orig_bytes": -1, "new_bytes": -1,
           "quality": "", "edge": "", "anim": "", "out_file": "", "status": ""}

    raw = download(url)
    if raw is None:
        row["status"] = "下载失败"
        return row
    row["orig_bytes"] = len(raw)

    try:
        img = Image.open(io.BytesIO(raw))
        is_anim = getattr(img, "is_animated", False) and getattr(img, "n_frames", 1) > 1
        row["anim"] = "动图" if is_anim else "静态"

        try:
            if is_anim:
                data, q, edge = compress_animated(img, MAX_EDGE)
            else:
                data, q, edge = compress_static(img, MAX_EDGE)
        except Exception as e:
            # 动图存 webp 失败等情况 -> 回退到首帧静态
            log(f"[WARN] {uid} 动图压缩失败，回退首帧静态: {e}")
            img.seek(0)
            data, q, edge = compress_static(img, MAX_EDGE)
            row["anim"] += "(回退静态)"

        out_path = os.path.join(OUT_DIR, f"{uid}.webp")
        with open(out_path, "wb") as f:
            f.write(data)

        row["new_bytes"] = len(data)
        row["quality"] = q
        row["edge"] = edge
        row["out_file"] = out_path
        row["status"] = "OK" if len(data) <= SIZE_LIMIT else "仍超2M(已尽力)"
    except Exception as e:
        row["status"] = f"压缩异常: {e}"
        log(f"[ERROR] {uid} {url} -> {e}")
    return row


def main():
    if not os.path.exists(SRC_LIST):
        print(f"[ERROR] 找不到列表文件: {SRC_LIST}\n请先运行 Test.py 生成 avatar_over_2m.xlsx")
        return
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = load_workbook(SRC_LIST)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 跳过表头，取 uid + avatar 两列
    tasks = [(r[0], r[1]) for r in rows[1:] if len(r) >= 2 and r[1]]
    total = len(tasks)
    print(f"待处理 {total} 张图片，输出到: {OUT_DIR}")

    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = [pool.submit(process, uid, url) for uid, url in tasks]
        for i, fut in enumerate(concurrent.futures.as_completed(futures), start=1):
            results.append(fut.result())
            if i % 50 == 0 or i == total:
                print(f"  进度 {i}/{total}")

    # 写报表
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "report"
    out_ws.append(["uid", "原URL", "原大小MB", "新大小MB", "质量", "长边px",
                   "类型", "输出文件", "状态"])
    ok = 0
    for r in results:
        if r["status"] == "OK":
            ok += 1
        out_ws.append([
            r["uid"], r["url"],
            round(r["orig_bytes"] / 1024 / 1024, 2) if r["orig_bytes"] > 0 else "",
            round(r["new_bytes"] / 1024 / 1024, 2) if r["new_bytes"] > 0 else "",
            r["quality"], r["edge"], r["anim"], r["out_file"], r["status"],
        ])
    out_wb.save(REPORT)

    print(f"\n完成: 成功压到2M内 {ok}/{total}")
    print(f"图片目录: {OUT_DIR}")
    print(f"报表: {REPORT}")


if __name__ == "__main__":
    main()
