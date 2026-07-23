"""
bao-guang.py  —  多平台播放量/浏览量抓取
支持：X (Twitter) · Instagram · Facebook · TikTok · Threads
各平台并行运行，各自使用独立浏览器 profile 保存登录状态。
"""

from playwright.sync_api import sync_playwright, BrowserContext, Page
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys
import time
import os
import re
import traceback
import threading
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ──────────────────────────────────────────────
# 配置：在此填写所有账号 URL（自动识别平台）
# ──────────────────────────────────────────────
URLS = [
]

# Google Sheets 配置（留空则跳过，直接用上面的 URLS）
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1FYmDs2lKsrKQ_cye7xpUguuYAY2TsMK57GsFtYXTBhE/edit?gid=0#gid=0"

# URL → 元数据 的映射，由 fetch_google_sheet_urls() 填充
# 每条值形如 {"seq": "1", "name": "王永洋", "tab": "Sheet1"}
URL_META: dict[str, dict] = {}

# 平台识别规则（域名关键词 → 平台名）
_PLATFORM_RULES = [
    ("x.com",        "X"),
    ("twitter.com",  "X"),
    ("instagram.com","Instagram"),
    ("facebook.com", "Facebook"),
    ("tiktok.com",   "TikTok"),
    ("threads.com",  "Threads"),
    ("threads.net",  "Threads"),
]

def detect_platform(url: str) -> str | None:
    for keyword, platform in _PLATFORM_RULES:
        if keyword in url:
            return platform
    return None

def group_urls_by_platform(urls: list[str]) -> dict[str, list[str]]:
    """按平台分组，忽略无法识别的 URL"""
    grouped: dict[str, list[str]] = {}
    for url in urls:
        platform = detect_platform(url)
        if platform is None:
            print(f"[警告] 无法识别平台，已跳过：{url}")
            continue
        grouped.setdefault(platform, []).append(url)
    return grouped

# ──────────────────────────────────────────────
# 浏览器 Profile 目录（各平台独立，互不干扰）
# ──────────────────────────────────────────────
# 打包为 exe 后 __file__ 指向临时解压目录，需改用 exe 所在目录
if getattr(sys, "frozen", False):
    _BASE = os.path.dirname(sys.executable)
else:
    _BASE = os.path.dirname(os.path.abspath(__file__))


def _find_chrome() -> str:
    """查找系统 Chrome 可执行文件路径，找不到则抛出异常"""
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe")
        path, _ = winreg.QueryValueEx(key, "")
        if os.path.exists(path):
            return path
    except Exception:
        pass
    candidates = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe"),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    raise RuntimeError(
        "未找到系统 Chrome，请先安装 Google Chrome：https://www.google.com/chrome/"
    )


CHROME_PATH = _find_chrome()


PROFILE = {
    "x":        os.path.join(_BASE, ".profile_x"),
    "instagram": os.path.join(_BASE, ".profile_instagram"),
    "facebook": os.path.join(_BASE, ".profile_facebook"),
    "threads":  os.path.join(_BASE, ".profile_threads"),
    "tiktok":   os.path.join(_BASE, ".profile_tiktok"),
}

# ── 持久化浏览器上下文（跨多次 main() 调用保持不关闭）──────────────
# platform -> (playwright, BrowserContext, thread_id)
_CTX_STORE: dict[str, tuple] = {}

# TikTok 的反爬能识别 Playwright 亲自启动的浏览器（表现为视频列表加载不出来，
# 而同一个 profile 手动打开 Chrome 却一切正常）。因此 TikTok 改成：
# 用和「登录」按钮完全相同的方式启动普通 Chrome，只多开一个调试端口，
# 再让 Playwright 通过 CDP 附加上去，浏览器本身没有任何自动化启动参数。
TIKTOK_CDP_PORT = 9222
# {"playwright":..., "browser":..., "context":...}
_TT_CDP: dict = {}
# Chrome 进程句柄，CDP 关不掉时用来兜底结束进程
_TT_PROC = None


def _is_port_open(port: int, host: str = "127.0.0.1") -> bool:
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        return s.connect_ex((host, port)) == 0


def _launch_plain_chrome(profile_dir: str, url: str, debug_port: int | None = None):
    """用普通 Chrome 打开（不带任何自动化标志），cookie 落在共用 profile 目录"""
    import subprocess
    args = [CHROME_PATH, f"--user-data-dir={profile_dir}",
            "--no-first-run", "--no-default-browser-check"]
    if debug_port:
        args.append(f"--remote-debugging-port={debug_port}")
    args.append(url)
    return subprocess.Popen(args)


# 命令行开标签后，留给页面自行加载、拿到 item_list 数据的时间（秒）
TIKTOK_PAGE_SETTLE = 6.0


def _tiktok_ensure_chrome():
    """确保带调试端口的 Chrome 已在运行（只启动，不附加）"""
    global _TT_PROC
    if _is_port_open(TIKTOK_CDP_PORT):
        return
    print(f"[TIKTOK] 启动 Chrome（调试端口 {TIKTOK_CDP_PORT}）...")
    _TT_PROC = _launch_plain_chrome(PROFILE["tiktok"], "about:blank", TIKTOK_CDP_PORT)
    deadline = time.time() + 30
    while time.time() < deadline:
        if _is_port_open(TIKTOK_CDP_PORT):
            return
        time.sleep(0.5)
    raise RuntimeError(
        f"Chrome 调试端口 {TIKTOK_CDP_PORT} 未就绪。"
        "若已有使用同一 profile 的 Chrome 窗口开着，请先全部关闭再重试。")


def _attach_tiktok() -> tuple:
    """附加到已在运行的 Chrome，返回 (playwright, context)"""
    if _TT_CDP:
        ctx = _TT_CDP.get("context")
        if ctx is not None and _is_context_alive(ctx):
            return _TT_CDP["playwright"], ctx
        _close_tiktok_cdp()

    p = sync_playwright().start()
    browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{TIKTOK_CDP_PORT}")
    ctx = browser.contexts[0] if browser.contexts else browser.new_context()
    _TT_CDP.update({"playwright": p, "browser": browser, "context": ctx})
    return p, ctx


def _tiktok_open_tab(url: str):
    """
    用命令行让已运行的 Chrome 在新标签页打开 URL。
    关键：此刻不能有 CDP 连接——webmssdk 在页面初始化时检测调试连接，
    一旦被发现就生成废签名（X-Bogus=1），item_list 会返回 200 + 空 body。
    """
    _launch_plain_chrome(PROFILE["tiktok"], url)


def _tiktok_click_sort_popular(page: Page) -> bool:
    """
    切换到「热门」排序，借此重新发起一次列表请求。
    默认的「最新」排序在自动化环境下可能拿不到数据（签名被判废，
    item_list 返回 200 + 空 body），而切排序能正常加载出来。
    抓播放量不关心视频顺序，用哪种排序都一样。
    """
    try:
        page.wait_for_selector("#user-post-sort-control button", timeout=10000)
    except Exception:
        return False
    btns = page.query_selector_all("#user-post-sort-control button")
    # aria-label 随界面语言变化，命中不了就退回按位置取（最新 \ 热门 \ 最旧）
    for b in btns:
        label = (b.get_attribute("aria-label") or "").strip()
        if label in ("热门", "熱門", "Popular"):
            b.click()
            return True
    if len(btns) >= 2:
        btns[1].click()
        return True
    return False


def _tiktok_find_page(ctx, url: str):
    """在已打开的标签页中找到目标 URL 对应的页面"""
    target = url.split("?")[0].rstrip("/").lower()
    for pg in ctx.pages:
        try:
            if pg.url.split("?")[0].rstrip("/").lower() == target:
                return pg
        except Exception:
            continue
    return None


def _close_tiktok_cdp():
    """
    只断开 CDP 连接，浏览器留着。
    抓每个账号之间都要断开一次（下一个账号的页面必须在无调试连接的状态下加载），
    整轮结束才调 _quit_tiktok_chrome() 真正退出浏览器。
    """
    browser = _TT_CDP.get("browser")
    p = _TT_CDP.get("playwright")
    _TT_CDP.clear()
    try:
        if browser is not None:
            browser.close()
    except Exception:
        pass
    try:
        if p is not None:
            p.stop()
    except Exception:
        pass


def _quit_tiktok_chrome():
    """关闭 TikTok 的 Chrome，与其它平台跑完即关闭浏览器的行为保持一致"""
    global _TT_PROC
    if not _TT_CDP and not _is_port_open(TIKTOK_CDP_PORT):
        _TT_PROC = None
        return
    # 优先走 CDP 让 Chrome 正常退出，这样 profile（含登录 cookie）能正确落盘
    try:
        if not _TT_CDP:
            _attach_tiktok()
        browser = _TT_CDP.get("browser")
        if browser is not None:
            browser.new_browser_cdp_session().send("Browser.close")
    except Exception:
        pass
    _close_tiktok_cdp()

    # 等 Chrome 退出；超时则直接结束进程兜底
    deadline = time.time() + 10
    while time.time() < deadline:
        if not _is_port_open(TIKTOK_CDP_PORT):
            _TT_PROC = None
            return
        time.sleep(0.5)
    if _TT_PROC is not None:
        try:
            if _TT_PROC.poll() is None:
                _TT_PROC.terminate()
        except Exception:
            pass
        _TT_PROC = None


def _is_context_alive(ctx) -> bool:
    try:
        _ = ctx.pages
        return True
    except Exception:
        return False


def _close_platform_context(platform: str):
    """关闭指定平台的浏览器上下文，必须在创建该 context 的同一线程内调用"""
    entry = _CTX_STORE.pop(platform, None)
    if entry is None:
        return
    p, ctx, _ = entry
    try:
        ctx.close()
    except Exception:
        pass
    try:
        p.stop()
    except Exception:
        pass


def _ensure_context(platform: str) -> tuple:
    """返回平台对应的 (playwright, context)；context 失效时才重建，存活则直接复用。"""
    current_tid = threading.get_ident()
    if platform in _CTX_STORE:
        p, ctx, _ = _CTX_STORE[platform]
        if _is_context_alive(ctx):
            # context 仍然存活（无论哪个线程创建的），直接复用并更新线程记录
            _CTX_STORE[platform] = (p, ctx, current_tid)
            return p, ctx
        # context 已失效，关闭后重建
        try:
            ctx.close()
        except Exception:
            pass
        try:
            p.stop()
        except Exception:
            pass
        del _CTX_STORE[platform]

    p = sync_playwright().start()
    ctx = make_context(p, platform)
    _CTX_STORE[platform] = (p, ctx, current_tid)
    return p, ctx


def _get_or_new_page(ctx):
    """复用 context 中第一个已有页面，没有则新建"""
    pages = ctx.pages
    if pages:
        pages[0].bring_to_front()
        return pages[0]
    return ctx.new_page()


def open_login_page(platform: str):
    """
    直接用 Chrome（非 Playwright 控制）打开登录页，避免被平台识别为自动化。
    登录成功后 cookie 保存在与 Playwright 共用的持久化 profile 目录中，
    后续定时运行时 Playwright 会自动读取这些 cookie。
    """
    import webbrowser

    _PLATFORM_HOME = {
        "x":         "https://x.com",
        "instagram": "https://www.instagram.com",
        "facebook":  "https://www.facebook.com",
        "tiktok":    "https://www.tiktok.com",
        "threads":   "https://www.threads.com",
    }
    url = _PLATFORM_HOME.get(platform)
    if not url:
        return

    profile_dir = PROFILE.get(platform)
    if profile_dir and CHROME_PATH:
        # 先释放 Playwright 对该 profile 目录的文件锁，否则普通 Chrome 会回退到临时 profile，
        # 导致登录 cookie 写不进共用目录
        if platform == "tiktok":
            _close_tiktok_cdp()
        else:
            _close_platform_context(platform)
        # 用普通 Chrome（不带 Playwright 自动化标志）打开，cookie 写入同一 profile。
        # TikTok 顺便带上调试端口，登录完不必关窗口，抓取时直接附加到这个浏览器。
        try:
            _launch_plain_chrome(
                profile_dir, url,
                TIKTOK_CDP_PORT if platform == "tiktok" else None)
            print(f"[{platform.upper()}] Chrome 已打开 {url}")
            print(f"[{platform.upper()}] 请完成登录，关闭浏览器后 cookie 会自动保存，下次自动化运行即可使用。")
            return
        except Exception as e:
            print(f"[{platform.upper()}] 直接启动 Chrome 失败，回退到系统浏览器: {e}")

    # 回退：用系统默认浏览器打开
    webbrowser.open(url)
    print(f"[{platform.upper()}] 已用系统浏览器打开 {url}，请手动完成登录。")

# Excel 写入锁（多平台并行时保护文件）
_excel_lock = threading.Lock()

# ──────────────────────────────────────────────
# 通用工具
# ──────────────────────────────────────────────

def parse_view_count(text: str) -> int | None:
    """解析播放量文本，支持：3,926 / 3.9K / 1.2M / 1.2万/萬 / 3亿/億"""
    # \xa0 是不换行空格（&nbsp;），Facebook 繁体页面常用
    text = text.strip().replace(",", "").replace(" ", "").replace("\xa0", "")
    if re.fullmatch(r'\d+', text):
        return int(text)
    m = re.fullmatch(r'([\d.]+)([KkMmBb])', text)
    if m:
        multiplier = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
        return int(float(m.group(1)) * multiplier[m.group(2).upper()])
    m = re.fullmatch(r'([\d.]+)[万萬]', text)  # 简体万 + 繁体萬
    if m:
        return int(float(m.group(1)) * 10_000)
    m = re.fullmatch(r'([\d.]+)[亿億]', text)  # 简体亿 + 繁体億
    if m:
        return int(float(m.group(1)) * 100_000_000)
    return None


def safe_goto(page: Page, url: str, wait_until: str = "load"):
    """
    跳转页面，忽略内部重定向引发的导航中断异常；网络错误时每秒重试最多60次。
    wait_until 可传 "domcontentloaded"/"commit"：SPA 站点的内容就绪判据是
    wait_for_selector 而非 load 事件，等 load（要等完所有图片/视频）纯属浪费。
    """
    deadline = time.time() + 60
    attempt = 0
    while True:
        try:
            page.goto(url, wait_until=wait_until, timeout=60000)  # type: ignore[arg-type]
            return
        except Exception as e:
            err_str = str(e)
            if "interrupted by another navigation" in err_str:
                page.wait_for_load_state("domcontentloaded", timeout=30000)
                return
            if "net::ERR_" in err_str and time.time() < deadline:
                attempt += 1
                print(f"[网络错误] 第 {attempt} 次重试（{url[:60]}）...")
                time.sleep(1)
                continue
            raise


_STEALTH_SCRIPT = """
    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
    Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
    Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN','zh','en']});
    window.chrome = {runtime: {}};
    const _query = window.navigator.permissions.query;
    window.navigator.permissions.query = (p) =>
        p.name === 'notifications'
            ? Promise.resolve({state: Notification.permission})
            : _query(p);
"""


def make_context(p, platform: str) -> BrowserContext:
    """创建持久化浏览器上下文"""
    print(f"[{platform.upper()}] CHROME_PATH={CHROME_PATH}")
    print(f"[{platform.upper()}] exists={os.path.exists(CHROME_PATH) if CHROME_PATH else False}")
    ctx = p.chromium.launch_persistent_context(
        user_data_dir=PROFILE[platform],
        executable_path=CHROME_PATH,
        headless=False,
        args=[
            "--disable-blink-features=AutomationControlled",
            # 多标签页并行加载时，后台标签会被 Chrome 节流导致 hydration 变慢
            "--disable-background-timer-throttling",
            "--disable-backgrounding-occluded-windows",
            "--disable-renderer-backgrounding",
        ],
        viewport={"width": 1280, "height": 900},  # type: ignore[arg-type]
        locale="zh-CN",
    )
    print(f"[{platform.upper()}] Chrome 启动成功")
    ctx.add_init_script(_STEALTH_SCRIPT)
    return ctx


def scroll_until_stable(page: Page, delay: float = 2.0):
    """滚动到底部触发懒加载，等待网络静默后再继续（用于非 Facebook 平台）"""
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(delay)
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass  # 社交平台后台轮询可能导致 networkidle 永远不触发，忽略超时


def scroll_and_wait_new_elements(page: Page, selector: str, current_dom_count: int,
                                  poll: float = 0.4, timeout: float = 3.0) -> int:
    """
    滚动后等待：新容器数量 > current_dom_count 即立即返回。
    只要求"出现新容器"，不再要求"所有容器都已填充文字"——后者在 FB 通用类名下
    几乎永远不成立，会导致每次滚动都空耗满 timeout。骨架未渲染的元素外层
    parse_view_count 会自动跳过，下一轮滚动会再次抓到。
    超时后直接返回当前数量，交给外层 no_new 计数判断是否到底。
    """
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    deadline = time.time() + timeout
    while time.time() < deadline:
        time.sleep(poll)
        total = page.evaluate(
            f"() => document.querySelectorAll('{selector}').length")
        # 有新容器即返回
        if total > current_dom_count:
            return total
    return len(page.query_selector_all(selector))


def _wait_for_content(page: Page, selector: str, timeout: int = 15000):
    """等待指定选择器出现，超时则静默继续"""
    try:
        page.wait_for_selector(selector, timeout=timeout)
    except Exception:
        pass


# ──────────────────────────────────────────────
# X (Twitter)
# ──────────────────────────────────────────────

def _x_is_logged_in(page: Page) -> bool:
    for selector in [
        '[data-testid="SideNav_AccountSwitcher_Button"]',
        '[data-testid="AppTabBar_Home_Link"]',
        'a[href="/home"]',
    ]:
        try:
            page.wait_for_selector(selector, timeout=3000)
            return True
        except Exception:
            pass
    return False


def _x_ensure_login(page: Page):
    safe_goto(page, "https://x.com/")
    if _x_is_logged_in(page):
        print("[X] 检测到已登录，直接开始抓取。")
    else:
        print("[X] 尚未登录，请在浏览器中完成登录后按 Enter 继续...")
        input()
        page.wait_for_load_state("load", timeout=60000)


def _x_extract_views(aria_label: str) -> int | None:
    m = re.search(r'([\d,]+)\s+views', aria_label)
    if m:
        return int(m.group(1).replace(",", ""))
    return None


def scrape_x(urls: list[str], check_login: bool = True) -> dict[str, list[dict]]:
    results: dict[str, list[dict]] = {}
    _, context = _ensure_context("x")
    page = _get_or_new_page(context)
    if check_login:
        _x_ensure_login(page)
    for url in urls:
        print(f"\n[X] 正在访问：{url}")
        try:
            safe_goto(page, url)
        except Exception as e:
            print(f"[X] 网络错误，跳过当前账号，已保留 {len(results)} 个结果：{e}")
            results[url] = []
            continue
        time.sleep(3)
        _wait_for_content(page, '[data-testid="tweet"]')

        seen: set[str] = set()
        items: list[dict] = []
        last_count = 0
        no_new = 0

        while True:
            for tweet in page.query_selector_all('[data-testid="tweet"]'):
                link_el = tweet.query_selector('a[href*="/status/"]')
                tid = link_el.get_attribute("href") if link_el else None
                if not tid or tid in seen:
                    continue
                seen.add(tid)

                # 过滤转发
                ctx = tweet.query_selector('[data-testid="socialContext"]')
                if ctx:
                    t = ctx.inner_text().strip().lower()
                    if "reposted" in t or "转发" in t:
                        continue

                # 获取浏览量
                views_el = tweet.query_selector('[aria-label*="views. View post analytics"]')
                if not views_el:
                    views_el = tweet.query_selector('[role="group"][aria-label*="views"]')
                if not views_el:
                    continue

                vc = _x_extract_views(views_el.get_attribute("aria-label") or "")
                if vc is None:
                    continue

                items.append({"index": len(items) + 1, "views": vc,
                              "url": url, "href": "https://x.com" + tid})

            cur = len(items)
            print(f"[X]   已收集 {cur} 条原创推文...")
            if cur == last_count:
                no_new += 1
                if no_new >= 3:
                    break
            else:
                no_new = 0
            last_count = cur
            scroll_until_stable(page)
            _wait_for_content(page, '[data-testid="tweet"]')

        results[url] = items
    _close_platform_context("x")
    return results


# ──────────────────────────────────────────────
# Instagram
# ──────────────────────────────────────────────

IG_VIEW_SELECTOR = (
    "span.xdj266r.x14z9mp.xat24cr.x1lziwak.xexx8yu"
    ".xyri2b.x18d9i69.x1c1uobl.x1hl2dhg.x16tdsg8.x1vvkbs"
)
IG_REEL_LINK_SELECTOR = 'a[href*="/reel/"]'


def _ig_is_logged_in(page: Page) -> bool:
    try:
        page.wait_for_selector('nav[role="navigation"]', timeout=5000)
        return page.query_selector('a[href="/accounts/login/"]') is None
    except Exception:
        return False


def _ig_has_login_dialog(page: Page) -> bool:
    """检测页面是否弹出了登录对话框（浏览时触发的登录墙）"""
    try:
        dialog = page.query_selector('[role="dialog"]')
        if not dialog:
            return False
        # 对话框内含登录链接
        if dialog.query_selector('a[href*="/accounts/login/"]'):
            return True
        # 对话框内文字含登录/注册关键词
        text = dialog.inner_text()
        return any(kw in text for kw in ["Log in", "登录", "Sign up", "注册"])
    except Exception:
        return False


def _ig_wait_for_login(page: Page):
    """检测到登录弹窗时，自动轮询直到真正登录完成才继续"""
    print("[IG] ⚠ 检测到登录弹窗，请在浏览器中完成登录，登录完成后将自动继续...")
    # 最多等待 10 分钟，每 2 秒检查一次
    for _ in range(300):
        time.sleep(2)
        try:
            cur_url = page.url
        except Exception:
            continue
        # 还在登录页或登录表单可见 → 继续等待
        if "login" in cur_url or "accounts/login" in cur_url:
            continue
        if page.query_selector('input[name="username"], input[name="password"]'):
            continue
        if _ig_has_login_dialog(page):
            continue
        print("[IG] ✅ 已确认登录完成，继续抓取。")
        return
    print("[IG] 等待登录超时，继续尝试抓取...")


def _ig_ensure_login(page: Page):
    safe_goto(page, "https://www.instagram.com/")
    if _ig_is_logged_in(page):
        print("[IG] 检测到已登录，直接开始抓取。")
    else:
        print("[IG] 尚未登录，请在浏览器中完成登录后按 Enter 继续...")
        input()


def scrape_instagram(urls: list[str], check_login: bool = True) -> dict[str, list[dict]]:
    results: dict[str, list[dict]] = {}
    _, context = _ensure_context("instagram")
    page = _get_or_new_page(context)
    if check_login:
        _ig_ensure_login(page)

    _js_extract = """
    (sel) => {
        const links = document.querySelectorAll('a[href*="/reel/"]');
        const out = [];
        for (const a of links) {
            const spans = a.querySelectorAll(sel);
            if (spans.length === 0) continue;
            const last = spans[spans.length - 1];
            out.push({href: a.href, text: (last.innerText || '').trim()});
        }
        return out;
    }
    """

    for url in urls:
        print(f"\n[IG] 正在访问：{url}")
        try:
            safe_goto(page, url)
        except Exception as e:
            print(f"[IG] 网络错误，跳过当前账号，已保留 {len(results)} 个结果：{e}")
            results[url] = []
            continue
        time.sleep(3)

        # 页面加载后立即检查是否弹出登录框
        if _ig_has_login_dialog(page):
            _ig_wait_for_login(page)

        _wait_for_content(page, IG_REEL_LINK_SELECTOR)

        items: list[dict] = []
        seen_hrefs: set[str] = set()
        last_new_time = time.time()
        IDLE_TIMEOUT = 10.0  # 连续 10 秒无新内容则认为到底

        while True:
            # 每次滚动前检查登录弹窗
            if _ig_has_login_dialog(page):
                _ig_wait_for_login(page)
                last_new_time = time.time()  # 登录等待期间不计入空闲

            try:
                reel_stats = page.evaluate(_js_extract, IG_VIEW_SELECTOR) or []
            except Exception:
                reel_stats = []

            for entry in reel_stats:
                href = entry.get("href", "")
                text = entry.get("text", "")
                if not href or href in seen_hrefs:
                    continue
                vc = parse_view_count(text)
                if vc is None:
                    continue
                seen_hrefs.add(href)
                items.append({"index": len(items) + 1, "views": vc, "url": url, "href": href})
                last_new_time = time.time()

            print(f"[IG]   已加载 {len(items)} 个播放量...")

            if time.time() - last_new_time > IDLE_TIMEOUT:
                print(f"[IG]   {IDLE_TIMEOUT:.0f}s 内无新内容，判断已到底。")
                break

            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1)

        results[url] = items
    _close_platform_context("instagram")
    return results


# ──────────────────────────────────────────────
# Facebook
# ──────────────────────────────────────────────

FB_VIEW_SELECTOR = "span.x1lliihq.x6ikm8r.x10wlt62.x1n2onr6.xlyipyv.xuxw1ft"


def _fb_is_logged_in(page: Page) -> bool:
    for label in ["你的主页", "Your profile", "Profile"]:
        try:
            page.wait_for_selector(f'[aria-label="{label}"]', timeout=2000)
            return True
        except Exception:
            pass
    return page.query_selector('a[href*="login"]') is None


def _fb_page_has_login_wall(page: Page) -> bool:
    """检测当前页面是否出现登录墙（含"登录"/"Log in"文字的遮罩）"""
    result = page.evaluate("""
        () => {
            // 检查页面是否含有登录遮罩：查找含"登录"或"Log in"的按钮/链接
            const texts = ["登录", "Log in", "Login"];
            for (const el of document.querySelectorAll('a[href*="login"], button')) {
                const t = (el.innerText || "").trim();
                if (texts.includes(t)) return true;
            }
            // 检查用户提供的具体 span 类名
            for (const span of document.querySelectorAll(
                'span.x1lliihq.x6ikm8r.x10wlt62.x1n2onr6.xlyipyv.xuxw1ft')) {
                const t = (span.innerText || "").trim();
                if (texts.includes(t)) return true;
            }
            return false;
        }
    """)
    return bool(result)


def _fb_ensure_login(page: Page):
    safe_goto(page, "https://www.facebook.com/")
    time.sleep(3)
    if _fb_is_logged_in(page):
        print("[FB] 检测到已登录，直接开始抓取。")
    else:
        print("[FB] 尚未登录，请在浏览器中完成登录后按 Enter 继续...")
        input()


def scrape_facebook(urls: list[str], check_login: bool = True) -> dict[str, list[dict]]:
    results: dict[str, list[dict]] = {}
    _, context = _ensure_context("facebook")
    page = _get_or_new_page(context)
    if check_login:
        _fb_ensure_login(page)

    for url in urls:
        print(f"\n[FB] 正在访问：{url}")
        try:
            safe_goto(page, url)
        except Exception as e:
            print(f"[Facebook] 网络错误，跳过当前账号，已保留 {len(results)} 个结果：{e}")
            results[url] = []
            continue
        time.sleep(4)
        _wait_for_content(page, FB_VIEW_SELECTOR, timeout=5000)

        if _fb_page_has_login_wall(page):
            print(f"[FB] 检测到登录墙，跳过抓取，曝光量记为 -1：{url}")
            results[url] = [{"index": 1, "views": -1, "url": url, "href": ""}]
            continue

        seen: set[str] = set()
        items: list[dict] = []
        last_count = 0
        no_new = 0
        dom_count = len(page.query_selector_all(FB_VIEW_SELECTOR))

        while True:
            data = page.evaluate(f"""
                () => {{
                    const spans = document.querySelectorAll("{FB_VIEW_SELECTOR}");
                    const out = [];
                    spans.forEach((span, i) => {{
                        const text = span.innerText.trim();
                        if (!text) return;
                        let el = span, reelHref = null;
                        for (let j = 0; j < 15; j++) {{
                            el = el.parentElement;
                            if (!el) break;
                            const a = el.querySelector('a[href*="/reel/"]');
                            if (a) {{ reelHref = a.href; break; }}
                            if (el.tagName === "A" && (el.href || "").includes("/reel/")) {{
                                reelHref = el.href; break;
                            }}
                        }}
                        out.push({{ i, text, reelHref }});
                    }});
                    return out;
                }}
            """)

            for item in data:
                vc = parse_view_count(item.get("text", ""))
                if vc is None:
                    continue
                href = item.get("reelHref") or ""
                # 无链接时按 DOM 位置去重，不能按文本——两条不同 Reel
                # 播放量文本可能相同，按文本会被误判成重复而丢数据
                key = href if href else f"idx:{item.get('i')}"
                if key in seen:
                    continue
                seen.add(key)
                items.append({"index": len(items) + 1, "views": vc,
                              "url": url, "href": href})

            cur = len(items)
            print(f"[FB]   已收集 {cur} 条 Reel...")
            if cur == last_count:
                no_new += 1
                if no_new >= 3:
                    break
            else:
                no_new = 0
            last_count = cur
            # 滚动后轮询 DOM 元素数，新元素一出现立即继续，无需等满 networkidle 超时
            dom_count = scroll_and_wait_new_elements(page, FB_VIEW_SELECTOR, dom_count)

        results[url] = items
    _close_platform_context("facebook")
    return results


# ──────────────────────────────────────────────
# TikTok
# ──────────────────────────────────────────────

def scrape_tiktok(urls: list[str], check_login: bool = True) -> dict[str, list[dict]]:  # noqa: check_login unused
    """
    TikTok 专用流程：页面必须由 Chrome 自己打开、加载完成后才允许附加 CDP。
    check_login 在这里是空的——未登录同样能拿到播放量，而登录检查会提前
    建立 CDP 连接，反而破坏上面这个时序。
    """
    results: dict[str, list[dict]] = {}
    _tiktok_ensure_chrome()

    for url in urls:
        print(f"\n[TT] 正在访问：{url}")
        page = None
        try:
            # 第一步：无 CDP 连接状态下让 Chrome 自行打开并加载完页面
            _close_tiktok_cdp()
            _tiktok_open_tab(url)
            time.sleep(TIKTOK_PAGE_SETTLE)

            # 第二步：页面已加载完，此时再附加
            _, context = _attach_tiktok()
            page = _tiktok_find_page(context, url)
            if page is None:
                raise RuntimeError(f"未找到已打开的标签页：{url}")
            page.bring_to_front()

            try:
                page.wait_for_selector('[data-e2e="video-views"]', timeout=15000)
            except Exception:
                print("[TT]   首屏没有播放量，切到「热门」排序重新触发加载...")
                if not _tiktok_click_sort_popular(page):
                    print("[TT]   没找到排序控件")
                page.wait_for_selector('[data-e2e="video-views"]', timeout=20000)

            seen: set[str] = set()
            items: list[dict] = []
            last_count = 0
            no_new = 0

            while True:
                # 每个播放量元素向上第 2 层即可取到自己那条视频链接，
                # 留 15 层余量防改版；带位置下标便于链接缺失时兜底去重。
                data = page.evaluate("""
                    () => {
                        const els = document.querySelectorAll('[data-e2e="video-views"]');
                        const out = [];
                        els.forEach((el, i) => {
                            const text = (el.innerText || "").trim();
                            let cur = el, href = null;
                            for (let j = 0; j < 15; j++) {
                                cur = cur.parentElement;
                                if (!cur) break;
                                const a = cur.querySelector('a[href*="/video/"]');
                                if (a) { href = a.href; break; }
                            }
                            out.push({ i, text, href });
                        });
                        return out;
                    }
                """)

                for entry in data:
                    vc = parse_view_count(entry.get("text", ""))
                    if vc is None:
                        continue  # 骨架屏未渲染，下一轮滚动会再抓到
                    href = entry.get("href") or ""
                    # 无链接时按 DOM 位置去重，不能按文本——两条不同视频
                    # 播放量文本可能相同，按文本会被误判成重复而丢数据
                    key = href if href else f"idx:{entry.get('i')}"
                    if key in seen:
                        continue
                    seen.add(key)
                    items.append({"index": len(items) + 1, "views": vc,
                                  "url": url, "href": href})

                cur = len(data)
                print(f"[TT]   已加载 {cur} 个视频，已收集 {len(items)} 条播放量...")
                if cur == last_count:
                    no_new += 1
                    if no_new >= 3:
                        break
                else:
                    no_new = 0
                last_count = cur
                scroll_until_stable(page)
                _wait_for_content(page, '[data-e2e="video-views"]')

            results[url] = items
        except Exception as e:
            # 原来一律报"网络错误"，但选择器超时、风控拦截也会走到这里，会误导排查
            print(f"[TT] 抓取失败，跳过当前账号，已保留 {len(results)} 个结果："
                  f"{type(e).__name__}: {e}")
            results[url] = []
        finally:
            # 关掉本轮标签页，避免账号多时标签堆积
            if page is not None:
                try:
                    page.close()
                except Exception:
                    pass

    _quit_tiktok_chrome()
    return results


# ──────────────────────────────────────────────
# Threads
# ──────────────────────────────────────────────
# 主页不显示浏览量，需先收集帖子链接再逐条进详情页提取。
# 该类名同时用于导航/用户名等文本，必须配合"次浏览/views"文本模式过滤。
THREADS_VIEW_SELECTOR = "span.x1lliihq.x193iq5w.x6ikm8r.x10wlt62.xlyipyv.xuxw1ft"
_THREADS_VIEW_RE = re.compile(r'(次浏览|次瀏覽|views?)\s*$', re.IGNORECASE)
_THREADS_POST_LINK = 'a[href*="/post/"]'
_THREADS_PATH_RE = re.compile(r'/(@[^/]+)/post/([^/]+)')

# 滚动收集的到底判据：连续这么多秒没有出现新帖子链接就结束
_THREADS_NO_NEW_SECONDS = 10.0

# 详情页并行标签数：N 个标签同时加载，再依次收割浏览量。
# 若遇到限流/页面加载不稳，调回 1 即退化为原来的串行流程。
THREADS_POST_CONCURRENCY = 4

# 浏览量只是一个文本 span，图片/视频/字体全部不需要，拦掉可省去详情页大半加载时间
_THREADS_BLOCK_TYPES = {"image", "media", "font"}


def _threads_block_heavy(route):
    try:
        if route.request.resource_type in _THREADS_BLOCK_TYPES:
            route.abort()
        else:
            route.continue_()
    except Exception:
        pass


def _threads_ensure_login(page: Page):
    safe_goto(page, "https://www.threads.com/", wait_until="domcontentloaded")
    time.sleep(3)
    if page.query_selector('a[href*="/login"]') is None:
        print("[TH] 检测到已登录，直接开始抓取。")
    else:
        print("[TH] 尚未登录，请在浏览器中完成登录后按 Enter 继续...")
        input()


# 找到浏览量文本的 JS（找不到返回 null）。等待与提取共用同一判据：
# THREADS_VIEW_SELECTOR 是通用类名，导航/用户名也命中，只等类名出现会在
# 浏览量尚未渲染时就返回，必须以"文本匹配次浏览/views"为准。
_THREADS_VIEW_FIND_JS = """
    () => {
        const pat = /(次浏览|次瀏覽|views?)\\s*$/i;
        const hit = (list) => {
            for (const span of list) {
                const t = (span.innerText || "").trim();
                if (t.length < 30 && pat.test(t)) return t;
            }
            return null;
        };
        // 第二个 hit 是兜底：类名可能随版本变化
        return hit(document.querySelectorAll("%s")) || hit(document.querySelectorAll("span"));
    }
""" % THREADS_VIEW_SELECTOR


def _threads_wait_views(page: Page, timeout: int = 15000):
    """等到浏览量文本真正渲染出来。polling 必须显式给毫秒数：
    默认的 raf 轮询在后台标签页会被暂停，并行加载时大部分标签都是后台的。"""
    try:
        page.wait_for_function(
            f"() => ({_THREADS_VIEW_FIND_JS})() !== null",
            timeout=timeout, polling=300)
    except Exception:
        pass


def _threads_extract_views(page: Page) -> int | None:
    """在帖子详情页提取浏览量（如 '3.3 万次浏览' / '1,234 views'），整页仅主帖一处"""
    text = page.evaluate(_THREADS_VIEW_FIND_JS)
    if not text:
        return None
    return parse_view_count(_THREADS_VIEW_RE.sub("", text).strip())


def _threads_post_paths(page: Page) -> list[tuple[str, str]]:
    """
    取当前页面上的帖子链接，返回 [(path, account)]，path 形如 /@user/post/xxx。
    去掉 query 并排除 /media 等子页。主页是虚拟滚动、DOM 节点会被回收重排，
    所以一切"是否还有新内容"的判断都要基于这里的 path 集合，不能用 DOM 节点数。
    """
    from urllib.parse import urlparse
    hrefs = page.evaluate(f"""
        () => Array.from(document.querySelectorAll('{_THREADS_POST_LINK}'))
                  .map(a => a.getAttribute('href'))
    """) or []
    out: list[tuple[str, str]] = []
    for href in hrefs:
        if not href:
            continue
        path = href.split("?")[0]
        if path.startswith("http"):
            path = urlparse(path).path
        pm = _THREADS_PATH_RE.fullmatch(path)
        if pm:
            out.append((path, pm.group(1)))
    return out


def _threads_scroll_step(page: Page, seen: set[str],
                         poll: float = 0.3, timeout: float = 4.0):
    """滚动一屏并等到出现 seen 之外的新帖子链接就返回，避免固定 sleep + networkidle 空耗"""
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    deadline = time.time() + timeout
    while time.time() < deadline:
        time.sleep(poll)
        try:
            if any(p not in seen for p, _ in _threads_post_paths(page)):
                return
        except Exception:
            return


def _threads_collect_views(context, post_paths: list[str], src_url: str) -> list[dict]:
    """
    多标签并行取浏览量：先给每个标签发起导航（commit 即返回，加载在后台并行进行），
    再依次等待各标签的浏览量 span 出现并提取。
    """
    items: list[dict] = []
    if not post_paths:
        return items

    lanes = max(1, min(THREADS_POST_CONCURRENCY, len(post_paths)))
    pages: list[Page] = [_get_or_new_page(context)]
    extra: list[Page] = []
    for _ in range(lanes - 1):
        pg = context.new_page()
        extra.append(pg)
        pages.append(pg)

    try:
        # 在 context 上注册，对批次内所有标签（含后续新建）统一生效
        context.route("**/*", _threads_block_heavy)

        for start in range(0, len(post_paths), lanes):
            batch = post_paths[start:start + lanes]
            loaded: list[tuple[Page, str]] = []
            # 阶段一：并行发起导航
            for pg, path in zip(pages, batch):
                post_url = "https://www.threads.com" + path
                try:
                    safe_goto(pg, post_url, wait_until="commit")
                    loaded.append((pg, post_url))
                except Exception as e:
                    print(f"[TH] 帖子打开失败，跳过：{post_url}（{e}）")
            # 阶段二：依次收割
            for pg, post_url in loaded:
                _threads_wait_views(pg)
                try:
                    vc = _threads_extract_views(pg)
                except Exception as e:
                    print(f"[TH] 帖子解析失败，跳过：{post_url}（{e}）")
                    continue
                if vc is None:
                    print(f"[TH]   未找到浏览量：{post_url}")
                    continue
                items.append({"index": len(items) + 1, "views": vc,
                              "url": src_url, "href": post_url})
                print(f"[TH]   {len(items):>3}: {vc:>12,} 次  {post_url}")
    finally:
        try:
            context.unroute("**/*", _threads_block_heavy)
        except Exception:
            pass
        for pg in extra:
            try:
                pg.close()
            except Exception:
                pass
    return items


def scrape_threads(urls: list[str], check_login: bool = True) -> dict[str, list[dict]]:
    results: dict[str, list[dict]] = {}
    _, context = _ensure_context("threads")
    page = _get_or_new_page(context)
    if check_login:
        _threads_ensure_login(page)

    for url in urls:
        print(f"\n[TH] 正在访问：{url}")
        try:
            safe_goto(page, url, wait_until="domcontentloaded")
        except Exception as e:
            print(f"[TH] 网络错误，跳过当前账号，已保留 {len(results)} 个结果：{e}")
            results[url] = []
            continue
        _wait_for_content(page, _THREADS_POST_LINK)

        # 从主页 URL 提取用户名，只收集该用户自己的帖子（转发的 href 是他人用户名）
        m = re.search(r'threads\.(?:com|net)/(@[^/?#]+)', url)
        account = m.group(1) if m else None

        # 第一步：滚动主页收集帖子链接
        post_paths: list[str] = []
        seen_paths: set[str] = set()
        # seen_any 含他人转发，作为"页面还在出新内容"的信号；只看 post_paths
        # 会在连续几屏都是转发时误判到底
        seen_any: set[str] = set()
        last_printed = -1
        last_progress = time.time()
        while True:
            for path, acc in _threads_post_paths(page):
                if path not in seen_any:
                    seen_any.add(path)
                    last_progress = time.time()
                if account and acc != account:
                    continue
                if path in seen_paths:
                    continue
                seen_paths.add(path)
                post_paths.append(path)

            if len(post_paths) != last_printed:
                print(f"[TH]   已收集 {len(post_paths)} 条帖子链接...")
                last_printed = len(post_paths)
            if time.time() - last_progress >= _THREADS_NO_NEW_SECONDS:
                print(f"[TH]   {_THREADS_NO_NEW_SECONDS:.0f} 秒无新内容，"
                      f"收集结束，共 {len(post_paths)} 条")
                break
            _threads_scroll_step(page, seen_any)

        # 第二步：多标签并行访问帖子详情页，提取浏览量
        results[url] = _threads_collect_views(context, post_paths, url)
    _close_platform_context("threads")
    return results


# ──────────────────────────────────────────────
# 工具：从 URL 提取账号名
# ──────────────────────────────────────────────

_SKIP_SEGMENTS = {"reels", "videos", "posts", "profile.php", ""}

def extract_account(url: str) -> str:
    # Facebook：从 id= 参数提取
    if "id=" in url:
        m = re.search(r'id=(\d+)', url)
        return m.group(1) if m else url
    # 取 path 各段，过滤掉已知非用户名的部分，取第一个有效段
    from urllib.parse import urlparse
    path_parts = [p.split("?")[0] for p in urlparse(url).path.strip("/").split("/")]
    for part in path_parts:
        if part and part not in _SKIP_SEGMENTS:
            return part
    return url


# ──────────────────────────────────────────────
# 汇总输出（终端）
# ──────────────────────────────────────────────

def print_results(platform: str, all_results: dict[str, list[dict]]):
    print(f"\n{'='*10} {platform} {'='*10}")
    for url, items in all_results.items():
        account = extract_account(url)
        total = sum(i["views"] for i in items)
        print(f"\n  账号: {account}  共 {len(items)} 条内容")
        for item in items:
            link = item.get("href") or item.get("url", "")
            print(f"    {item['index']:>4}: {item['views']:>12,} 次  {link}")
        print(f"  合计: {total:,}")


# ──────────────────────────────────────────────
# Google Sheets 读取
# ──────────────────────────────────────────────

_PLATFORM_KW = ["x.com", "twitter.com", "instagram.com",
                "facebook.com", "tiktok.com", "youtube.com",
                "threads.com", "threads.net"]


def _read_xlsx(xlsx_path: str, label: str = "Sheet") -> dict[str, dict]:
    """
    读取 Excel 文件，按行解析，识别"序号"/"sheet名称"/"链接"列。
    直接解析 xlsx zip 内的 XML，完全绕开 openpyxl 样式解析。
    返回 {url: {"seq": str, "name": str, "tab": str}} 映射。
    """
    import zipfile
    import xml.etree.ElementTree as ET

    _SS  = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    _REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    _PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
    _url_re = re.compile(r'https?://\S+')

    def _col_idx(ref: str) -> int:
        """将列字母 'A'/'AB' 转为 0-based 整数索引"""
        m = re.match(r'([A-Za-z]+)', ref)
        if not m:
            return 0
        idx = 0
        for ch in m.group(1).upper():
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    def _cell_text(c_el, shared: list[str]) -> str:
        v_el = c_el.find(f"{{{_SS}}}v")
        if v_el is None or not v_el.text:
            return ""
        if c_el.get("t") == "s":
            try:
                return shared[int(v_el.text)]
            except (IndexError, ValueError):
                return ""
        return v_el.text.strip()

    # 表头关键词映射
    _SEQ_KW  = {"序号", "no", "no.", "#"}
    _NAME_KW = {"姓名", "sheet名称", "名称", "账号", "name"}
    _URL_KW  = {"链接", "主页地址", "url", "网址", "地址"}
    _TMPL_KW = {"模版", "模板", "template"}

    url_meta: dict[str, dict] = {}
    _tmpl_col_map: dict[str, int | None] = {}  # sheet -> tmpl col idx

    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            names = set(zf.namelist())

            # 共享字符串表
            shared: list[str] = []
            if "xl/sharedStrings.xml" in names:
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for si in root.findall(f".//{{{_SS}}}si"):
                    parts = [t.text for t in si.findall(f".//{{{_SS}}}t") if t.text]
                    shared.append("".join(parts))

            # sheet tab 名称
            rid_to_name: dict[str, str] = {}
            if "xl/workbook.xml" in names:
                root = ET.fromstring(zf.read("xl/workbook.xml"))
                for sh in root.findall(f".//{{{_SS}}}sheet"):
                    rid = sh.get(f"{{{_REL}}}id", "")
                    rid_to_name[rid] = sh.get("name", rid)

            rid_to_path: dict[str, str] = {}
            if "xl/_rels/workbook.xml.rels" in names:
                root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                for rel in root.findall(f".//{{{_PKG}}}Relationship"):
                    rid_to_path[rel.get("Id", "")] = rel.get("Target", "")

            for rid, sname in rid_to_name.items():
                target = rid_to_path.get(rid, "")
                if not target:
                    continue
                sheet_path = target if target in names else f"xl/{target}"
                if sheet_path not in names:
                    continue

                root = ET.fromstring(zf.read(sheet_path))

                # 解析每行：{row_num: {col_idx: text}}
                rows: dict[int, dict[int, str]] = {}
                for row_el in root.findall(f".//{{{_SS}}}row"):
                    r = int(row_el.get("r", 0))
                    row_data: dict[int, str] = {}
                    for c_el in row_el.findall(f"{{{_SS}}}c"):
                        cidx = _col_idx(c_el.get("r", ""))
                        text = _cell_text(c_el, shared)
                        if text:
                            row_data[cidx] = text
                    if row_data:
                        rows[r] = row_data

                # 在前 10 行中找表头
                seq_col = name_col = url_col = tmpl_col = None
                header_row = 0
                for r in sorted(rows)[:10]:
                    for cidx, val in rows[r].items():
                        v = val.strip().lower()
                        if v in _SEQ_KW:
                            seq_col = cidx
                        elif v in _NAME_KW:
                            name_col = cidx
                        elif v in _URL_KW:
                            url_col = cidx
                        elif v in _TMPL_KW:
                            tmpl_col = cidx
                    if seq_col is not None or url_col is not None:
                        header_row = r
                        break

                count = 0
                for r in sorted(rows):
                    if r <= header_row:
                        continue
                    row_data = rows[r]

                    # 找 URL：优先指定列，否则全行扫描
                    url = ""
                    search_cells = (
                        [row_data[url_col]] if url_col is not None and url_col in row_data
                        else list(row_data.values())
                    )
                    for text in search_cells:
                        for u in _url_re.findall(text):
                            u = u.rstrip(".,;:!?）】》」'\"")
                            if any(k in u for k in _PLATFORM_KW):
                                url = u
                                break
                        if url:
                            break

                    if not url or url in url_meta:
                        continue

                    seq  = row_data.get(seq_col, "") if seq_col is not None else ""
                    name = row_data.get(name_col, "") if name_col is not None else ""
                    tmpl = row_data.get(tmpl_col, "") if tmpl_col is not None else ""
                    url_meta[url] = {"seq": seq, "name": name, "tab": sname, "tmpl": tmpl}
                    count += 1

                print(f"[{label}] Sheet '{sname}' → {count} 条新 URL")

    except Exception as e:
        print(f"[{label}] 读取 xlsx 失败: {e}")
    finally:
        try:
            os.unlink(xlsx_path)
        except Exception:
            pass

    return url_meta


def fetch_google_sheet_urls() -> dict[str, dict]:
    """
    下载 Google Sheets 并解析平台 URL。支持两种格式：
    - 公开发布链接（…/pub?output=xlsx）：直接 HTTP 下载，无需浏览器
    - 私有文档链接（…/edit…）：打开浏览器登录后下载
    """
    import tempfile
    import urllib.request

    if not GOOGLE_SHEET_URL:
        return {}

    tmp_path = os.path.join(tempfile.gettempdir(),
                            f"google_sheet_{int(time.time())}.xlsx")

    # ── 判断是否为公开发布链接 ──────────────────────────────────────
    is_public = (
        "/pub?" in GOOGLE_SHEET_URL
        or "/pub#" in GOOGLE_SHEET_URL
        or GOOGLE_SHEET_URL.rstrip("/").endswith("/pub")
        or "/d/e/" in GOOGLE_SHEET_URL
    )

    if is_public:
        # 确保 output=xlsx 参数存在
        if "output=xlsx" in GOOGLE_SHEET_URL:
            download_url = GOOGLE_SHEET_URL
        elif "?" in GOOGLE_SHEET_URL:
            download_url = GOOGLE_SHEET_URL + "&output=xlsx"
        else:
            download_url = GOOGLE_SHEET_URL + "?output=xlsx"

        print(f"[Google] 检测到公开发布链接，直接下载...")
        try:
            urllib.request.urlretrieve(download_url, tmp_path)
            print(f"[Google] xlsx 已下载: {tmp_path}")
        except Exception as e:
            print(f"[Google] 下载失败: {e}")
            return {}

    else:
        # 私有文档：用浏览器登录后下载
        m = re.search(r'/spreadsheets/d/([^/]+)', GOOGLE_SHEET_URL)
        if not m:
            print("[Google] 无法从 URL 提取 Sheet ID")
            return {}
        sheet_id = m.group(1)
        export_url = (f"https://docs.google.com/spreadsheets/d/"
                      f"{sheet_id}/export?format=xlsx")
        profile_dir = os.path.join(_BASE, ".profile_google")

        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                executable_path=CHROME_PATH,
                headless=False,
                args=["--disable-blink-features=AutomationControlled"],
                viewport={"width": 1280, "height": 900},  # type: ignore[arg-type]
                locale="zh-CN",
            )
            page = context.new_page()

            print("[Google] 打开 Google Sheets，如需登录请在浏览器中完成...")
            safe_goto(page, GOOGLE_SHEET_URL)

            try:
                page.wait_for_selector(
                    "#docs-editor, .waffle-name-box, .grid-container",
                    timeout=600000)
                print("[Google] 文档已加载完成。")
            except Exception:
                print("[Google] 等待超时，尝试继续...")
            time.sleep(2)

            print("[Google] 开始下载 xlsx...")
            try:
                with page.expect_download(timeout=30000) as dl_info:
                    try:
                        page.goto(export_url, timeout=10000)
                    except Exception:
                        pass
                dl_info.value.save_as(tmp_path)
                print(f"[Google] xlsx 已下载: {tmp_path}")
            except Exception as e:
                print(f"[Google] 下载失败: {e}")
                context.close()
                return {}

            context.close()

    url_meta = _read_xlsx(tmp_path, label="Google")
    print(f"[Google] 共读取 {len(url_meta)} 条 URL")
    return url_meta


# ──────────────────────────────────────────────
# 导出 Excel（增量追加，按日期归档）
# ──────────────────────────────────────────────

_EXCEL_PATH: str = ""


def get_today_excel_path() -> str:
    global _EXCEL_PATH
    if not _EXCEL_PATH:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        _EXCEL_PATH = os.path.join(desktop, f"播放量汇总_{ts}.xlsx")
    return _EXCEL_PATH


def export_excel(all_platform_results: dict[str, dict]):
    """按 Google 文档原始顺序一次性写入 Excel"""
    # 展平为 url → items
    url_to_items: dict[str, list[dict]] = {}
    for results in all_platform_results.values():
        url_to_items.update(results)

    if not url_to_items:
        print("[Excel] 无数据可写入")
        return

    # 按原始 URLS 顺序排列，未在 URLS 中的补到末尾
    seen: set[str] = set()
    ordered: list[str] = []
    for u in URLS:
        if u in url_to_items and u not in seen:
            ordered.append(u)
            seen.add(u)
    for u in url_to_items:
        if u not in seen:
            ordered.append(u)

    path = get_today_excel_path()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    center      = Alignment(horizontal="center", vertical="center")

    # 构建 profile_url → platform 映射
    url_to_platform: dict[str, str] = {}
    for platform_key, results in all_platform_results.items():
        for u in results:
            url_to_platform[u] = platform_key

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    detail_headers = ["姓名", "唯一id", "平台", "模板", "播放量"]
    ws.append(detail_headers)
    for col_idx in range(1, len(detail_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    total_views = 0
    for url in ordered:
        items    = url_to_items[url]
        meta     = URL_META.get(url, {})
        name     = meta.get("name", "") or extract_account(url)
        tmpl     = meta.get("tmpl", "")
        platform = url_to_platform.get(url, detect_platform(url) or "")
        for item in items:
            unique_id = item.get("href") or item.get("url", "")
            ws.append([name, unique_id, platform, tmpl, item["views"]])
            total_views += item["views"]

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(path)
    print(f"\n[Excel] 共 {len(ordered)} 个账号，总曝光量 {total_views:,}")
    print(f"[Excel] 数据文件：{path}")


# ──────────────────────────────────────────────
# 主入口：并行抓取四个平台
# ──────────────────────────────────────────────

def main(check_login: bool = False, test_mode: bool = False):
    global URLS, URL_META, _EXCEL_PATH
    _EXCEL_PATH = ""  # 每次运行重置，强制生成新文件名

    # 从 Google Sheets 读取 URL 数据源
    if GOOGLE_SHEET_URL:
        sheet_map = fetch_google_sheet_urls()
        if sheet_map:
            URL_META = sheet_map
            URLS = list(sheet_map.keys())
            print(f"\n[Google] 共读取 {len(URLS)} 条 URL，开始抓取...\n")
        else:
            print("[Google] 未从表格读取到有效 URL，使用配置中的静态 URLS")

    scrape_fn = {
        "X":         scrape_x,
        "Instagram": scrape_instagram,
        "Facebook":  scrape_facebook,
        "TikTok":    scrape_tiktok,
        "Threads":   scrape_threads,
    }

    grouped = group_urls_by_platform(URLS)
    if test_mode:
        grouped = {p: urls[:3] for p, urls in grouped.items()}
        print("[测试模式] 每个平台最多处理 3 条 URL\n")

    tasks = {
        platform: (scrape_fn[platform], urls)
        for platform, urls in grouped.items()
        if platform in scrape_fn
    }

    if not tasks:
        print("[警告] 没有可识别的平台 URL，跳过抓取。")
        export_excel({})
        return

    all_platform_results: dict[str, dict] = {}

    with ThreadPoolExecutor(max_workers=len(tasks)) as executor:
        futures = {
            executor.submit(fn, urls, check_login): platform
            for platform, (fn, urls) in tasks.items()
        }
        for future in as_completed(futures):
            platform = futures[future]
            try:
                all_platform_results[platform] = future.result()
            except Exception as e:
                print(f"\n[{platform}] 抓取失败: {e}")
                all_platform_results[platform] = {}

    print("\n\n" + "=" * 50)
    print("  汇总结果")
    print("=" * 50)
    for platform, results in all_platform_results.items():
        print_results(platform, results)

    export_excel(all_platform_results)


def _close_all_contexts():
    """关闭所有平台的浏览器上下文"""
    _quit_tiktok_chrome()
    for platform, entry in list(_CTX_STORE.items()):
        p, ctx, _ = entry
        try:
            ctx.close()
        except Exception:
            pass
        try:
            p.stop()
        except Exception:
            pass
    _CTX_STORE.clear()


def _wait_until_next_8am():
    """阻塞到明天 08:00:00"""
    now = datetime.now()
    next_run = now.replace(hour=8, minute=0, second=0, microsecond=0)
    if now >= next_run:
        # 今天 8 点已过，等到明天
        from datetime import timedelta
        next_run += timedelta(days=1)
    wait_sec = (next_run - now).total_seconds()
    print(f"\n下次运行时间：{next_run.strftime('%Y-%m-%d %H:%M:%S')}，等待 {wait_sec/3600:.1f} 小时...")
    time.sleep(wait_sec)


if __name__ == "__main__":
    log_path = os.path.join(_BASE, "bao-guang.log")
    while True:
        try:
            main()
        except KeyboardInterrupt:
            raise
        except Exception:
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(traceback.format_exc())
            print(f"\n运行出错，详情已写入：{log_path}")
        try:
            _wait_until_next_8am()
        except KeyboardInterrupt:
            break
